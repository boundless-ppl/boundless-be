#!/usr/bin/env python3

import argparse
import datetime as dt
import sys
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


HEADER_ROW = 2
DATA_START_ROW = 5
UUID_NAMESPACE = uuid.UUID("8dc0d8af-1570-4e67-a7a4-4f2d9c12c8f3")


@dataclass
class CountryRow:
    negara_id: str
    nama: str
    nama_lokal: str
    benua: str
    mata_uang: str
    bahasa_resmi: str


@dataclass
class UniversityRow:
    universitas_id: str
    negara_id: str
    nama: str
    kota: str
    tipe: str
    deskripsi: str
    website: str
    ranking: str


@dataclass
class ProgramRow:
    program_id: str
    universitas_id: str
    nama_univ: str
    nama: str
    jenjang: str
    bahasa: str
    program_url: str


@dataclass
class CatalogRow:
    raw_id: str
    key: str


def normalize_header(value: object) -> str:
    text = str(value or "").strip()
    return text.replace(" 🔒", "")


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, dt.datetime):
        return value.isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def quote_sql(value: object) -> str:
    if value is None:
        return "NULL"
    text = str(value)
    return "'" + text.replace("'", "''") + "'"


def sql_text(value: str) -> str:
    value = value.strip()
    return "NULL" if value == "" else quote_sql(value)


def stable_uuid(kind: str, raw_id: str) -> str:
    return str(uuid.uuid5(UUID_NAMESPACE, f"{kind}:{raw_id.strip()}"))


def sql_bool(value: str) -> str:
    normalized = value.strip().upper()
    if normalized in {"TRUE", "T", "1", "YES"}:
        return "TRUE"
    if normalized in {"FALSE", "F", "0", "NO"}:
        return "FALSE"
    raise ValueError(f"invalid boolean value: {value!r}")


def sql_int(value: str) -> str:
    if value.strip() == "":
        raise ValueError("expected integer but got empty value")
    return str(int(float(value)))


def sql_date(value: str) -> str:
    value = value.strip()
    if value == "":
        return "NULL"
    if "T" in value:
        return f"TIMESTAMPTZ {quote_sql(value if value.endswith('Z') else value + 'Z')}"
    return f"TIMESTAMPTZ {quote_sql(value + 'T00:00:00Z')}"


def has_expected_prefix(value: str, prefix: str) -> bool:
    return value.strip().upper().startswith(prefix.upper())


def is_legend_row(value: str) -> bool:
    return value.strip().upper() == "LEGEND:"


def iter_sheet_rows(workbook, sheet_name: str) -> tuple[list[str], Iterable[tuple[int, dict[str, str]]]]:
    sheet = workbook[sheet_name]
    headers = [normalize_header(cell.value) for cell in sheet[HEADER_ROW]]

    def iterator():
        for row_number, raw_values in enumerate(
            sheet.iter_rows(min_row=DATA_START_ROW, values_only=True),
            start=DATA_START_ROW,
        ):
            values = [normalize_cell(cell) for cell in raw_values]
            if not any(values):
                continue
            yield row_number, {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}

    return headers, iterator()


def require_columns(sheet_name: str, headers: list[str], expected: list[str]) -> None:
    missing = [column for column in expected if column not in headers]
    if missing:
        raise ValueError(f"{sheet_name}: missing expected columns {missing}")


def collect_programs(workbook) -> dict[str, ProgramRow]:
    headers, rows = iter_sheet_rows(workbook, "Program")
    require_columns("Program", headers, [
        "program_id", "universitas_id", "nama_univ", "nama", "jenjang", "bahasa", "program_url",
    ])

    programs: dict[str, ProgramRow] = {}
    for row_number, row in rows:
        program_id = row["program_id"]
        if program_id == "":
            continue
        if (
            row["universitas_id"] == ""
            or row["nama_univ"] == ""
            or row["nama"] == ""
            or row["jenjang"] == ""
            or row["bahasa"] == ""
        ):
            continue
        programs[program_id] = ProgramRow(
            program_id=program_id,
            universitas_id=row["universitas_id"],
            nama_univ=row["nama_univ"],
            nama=row["nama"],
            jenjang=row["jenjang"],
            bahasa=row["bahasa"],
            program_url=row["program_url"],
        )
    return programs


def collect_countries(workbook) -> dict[str, CountryRow]:
    headers, rows = iter_sheet_rows(workbook, "Negara")
    require_columns("Negara", headers, [
        "negara_id", "nama", "nama_lokal", "benua", "mata_uang", "bahasa_resmi",
    ])

    countries: dict[str, CountryRow] = {}
    for _, row in rows:
        negara_id = row["negara_id"]
        if negara_id == "":
            continue
        countries[negara_id] = CountryRow(
            negara_id=negara_id,
            nama=row["nama"],
            nama_lokal=row["nama_lokal"],
            benua=row["benua"],
            mata_uang=row["mata_uang"],
            bahasa_resmi=row["bahasa_resmi"],
        )
    return countries


def collect_universities(workbook, countries: dict[str, CountryRow]) -> dict[str, UniversityRow]:
    headers, rows = iter_sheet_rows(workbook, "Universitas")
    require_columns("Universitas", headers, [
        "universitas_id", "negara_id", "nama", "kota", "tipe", "deskripsi", "website", "ranking",
    ])

    universities: dict[str, UniversityRow] = {}
    for row_number, row in rows:
        universitas_id = row["universitas_id"]
        if universitas_id == "":
            continue
        if row["negara_id"] == "" or row["nama"] == "" or row["kota"] == "" or row["tipe"] == "":
            continue
        if row["negara_id"] not in countries:
            raise ValueError(f"Universitas row {row_number}: unknown negara_id {row['negara_id']}")
        universities[universitas_id] = UniversityRow(
            universitas_id=universitas_id,
            negara_id=row["negara_id"],
            nama=row["nama"],
            kota=row["kota"],
            tipe=row["tipe"],
            deskripsi=row["deskripsi"],
            website=row["website"],
            ranking=row["ranking"],
        )
    return universities


def build_countries_sql(countries: dict[str, CountryRow]) -> list[str]:
    statements = []
    for row in countries.values():
        statements.append(
            "INSERT INTO countries (negara_id, nama, nama_lokal, benua, mata_uang, bahasa_resmi)\n"
            f"VALUES ({quote_sql(row.negara_id)}, {quote_sql(row.nama)}, {quote_sql(row.nama_lokal)}, {quote_sql(row.benua)}, {quote_sql(row.mata_uang)}, {quote_sql(row.bahasa_resmi)})\n"
            "ON CONFLICT (negara_id) DO UPDATE\n"
            "SET nama = EXCLUDED.nama,\n"
            "    nama_lokal = EXCLUDED.nama_lokal,\n"
            "    benua = EXCLUDED.benua,\n"
            "    mata_uang = EXCLUDED.mata_uang,\n"
            "    bahasa_resmi = EXCLUDED.bahasa_resmi,\n"
            "    updated_at = NOW();"
        )
    return statements


def build_universities_sql(universities: dict[str, UniversityRow]) -> list[str]:
    statements = []
    for row in universities.values():
        ranking_sql = "NULL" if row.ranking.strip() == "" else sql_int(row.ranking)
        statements.append(
            "INSERT INTO universities (id, external_id, negara_id, nama, kota, tipe, deskripsi, website, ranking)\n"
            f"VALUES ({quote_sql(stable_uuid('universities', row.universitas_id))}, {quote_sql(row.universitas_id)}, {quote_sql(row.negara_id)}, {quote_sql(row.nama)}, {quote_sql(row.kota)}, {quote_sql(row.tipe)}, {sql_text(row.deskripsi)}, {sql_text(row.website)}, {ranking_sql})\n"
            "ON CONFLICT (id) DO UPDATE\n"
            "SET external_id = EXCLUDED.external_id,\n"
            "    negara_id = EXCLUDED.negara_id,\n"
            "    nama = EXCLUDED.nama,\n"
            "    kota = EXCLUDED.kota,\n"
            "    tipe = EXCLUDED.tipe,\n"
            "    deskripsi = EXCLUDED.deskripsi,\n"
            "    website = EXCLUDED.website,\n"
            "    ranking = EXCLUDED.ranking,\n"
            "    updated_at = NOW();"
        )
    return statements


def build_programs_sql(programs: dict[str, ProgramRow], universities: dict[str, UniversityRow]) -> list[str]:
    statements = []
    for row in programs.values():
        if row.universitas_id == "" or row.nama == "" or row.jenjang == "" or row.bahasa == "":
            raise ValueError(f"Program {row.program_id}: required value missing")
        if row.universitas_id not in universities:
            raise ValueError(f"Program {row.program_id}: unknown universitas_id {row.universitas_id}")
        statements.append(
            "INSERT INTO programs (program_id, university_id, nama_univ, nama, jenjang, bahasa, program_url)\n"
            f"VALUES ({quote_sql(row.program_id)}, {quote_sql(stable_uuid('universities', row.universitas_id))}, {quote_sql(row.nama_univ)}, {quote_sql(row.nama)}, {quote_sql(row.jenjang)}, {quote_sql(row.bahasa)}, {quote_sql(row.program_url)})\n"
            "ON CONFLICT (program_id) DO UPDATE\n"
            "SET university_id = EXCLUDED.university_id,\n"
            "    nama_univ = EXCLUDED.nama_univ,\n"
            "    nama = EXCLUDED.nama,\n"
            "    jenjang = EXCLUDED.jenjang,\n"
            "    bahasa = EXCLUDED.bahasa,\n"
            "    program_url = EXCLUDED.program_url,\n"
            "    updated_at = NOW();"
        )
    return statements


def collect_catalog_rows(workbook, sheet_name: str, id_column: str) -> dict[str, CatalogRow]:
    headers, rows = iter_sheet_rows(workbook, sheet_name)
    require_columns(sheet_name, headers, [id_column, "key"])

    result: dict[str, CatalogRow] = {}
    for _, row in rows:
        raw_id = row[id_column]
        key = row["key"]
        if raw_id == "" or key == "":
            continue
        result[raw_id] = CatalogRow(raw_id=raw_id, key=key)
    return result


def build_requirement_catalog_sql(workbook) -> list[str]:
    headers, rows = iter_sheet_rows(workbook, "RequirementCatalog")
    require_columns("RequirementCatalog", headers, ["req_catalog_id", "key", "label", "kategori", "deskripsi"])

    statements = []
    for row_number, row in rows:
        if row["req_catalog_id"] == "" or not has_expected_prefix(row["req_catalog_id"], "REQ-"):
            continue
        if row["key"] == "" or row["label"] == "" or row["kategori"] == "":
            continue
        statements.append(
            "INSERT INTO requirement_catalog (req_catalog_id, key, label, kategori, deskripsi)\n"
            f"VALUES ({quote_sql(stable_uuid('requirement_catalog_key', row['key']))}, {quote_sql(row['key'])}, {quote_sql(row['label'])}, {quote_sql(row['kategori'])}, {sql_text(row['deskripsi'])})\n"
            "ON CONFLICT (key) DO UPDATE\n"
            "SET label = EXCLUDED.label,\n"
            "    kategori = EXCLUDED.kategori,\n"
            "    deskripsi = EXCLUDED.deskripsi;"
        )
    return statements


def build_benefit_catalog_sql(workbook) -> list[str]:
    headers, rows = iter_sheet_rows(workbook, "BenefitCatalog")
    require_columns("BenefitCatalog", headers, ["benefit_id", "key", "label", "kategori", "deskripsi"])

    statements = []
    for row_number, row in rows:
        if row["benefit_id"] == "" or not has_expected_prefix(row["benefit_id"], "BEN-"):
            continue
        if row["key"] == "" or row["label"] == "" or row["kategori"] == "":
            continue
        statements.append(
            "INSERT INTO benefit_catalog (benefit_id, key, label, kategori, deskripsi)\n"
            f"VALUES ({quote_sql(stable_uuid('benefit_catalog_key', row['key']))}, {quote_sql(row['key'])}, {quote_sql(row['label'])}, {quote_sql(row['kategori'])}, {sql_text(row['deskripsi'])})\n"
            "ON CONFLICT (key) DO UPDATE\n"
            "SET label = EXCLUDED.label,\n"
            "    kategori = EXCLUDED.kategori,\n"
            "    deskripsi = EXCLUDED.deskripsi;"
        )
    return statements


def build_funding_options_sql(workbook) -> list[str]:
    headers, rows = iter_sheet_rows(workbook, "FundingOptions")
    require_columns("FundingOptions", headers, ["funding_id", "nama_beasiswa", "deskripsi", "provider", "tipe_pembiayaan", "website"])

    statements = []
    for row_number, row in rows:
        if row["funding_id"] == "" or not has_expected_prefix(row["funding_id"], "FUND-"):
            continue
        if row["nama_beasiswa"] == "" or row["tipe_pembiayaan"] == "":
            continue
        statements.append(
            "INSERT INTO funding_options (funding_id, nama_beasiswa, deskripsi, provider, tipe_pembiayaan, website)\n"
            f"VALUES ({quote_sql(stable_uuid('funding_options', row['funding_id']))}, {quote_sql(row['nama_beasiswa'])}, {sql_text(row['deskripsi'])}, {quote_sql(row['provider'])}, {quote_sql(row['tipe_pembiayaan'])}, {quote_sql(row['website'])})\n"
            "ON CONFLICT (funding_id) DO UPDATE\n"
            "SET nama_beasiswa = EXCLUDED.nama_beasiswa,\n"
            "    deskripsi = EXCLUDED.deskripsi,\n"
            "    provider = EXCLUDED.provider,\n"
            "    tipe_pembiayaan = EXCLUDED.tipe_pembiayaan,\n"
            "    website = EXCLUDED.website;"
        )
    return statements


def build_admission_paths_sql(workbook, programs: dict[str, ProgramRow]) -> list[str]:
    headers, rows = iter_sheet_rows(workbook, "AdmissionPaths")
    require_columns("AdmissionPaths", headers, ["admission_id", "program_id", "nama_program", "nama", "intake", "deadline", "requires_supervisor", "website_url"])

    statements = []
    for row_number, row in rows:
        if row["admission_id"] == "" or is_legend_row(row["admission_id"]):
            continue
        if row["program_id"] == "" or row["nama"] == "" or row["intake"] == "" or row["deadline"] == "":
            raise ValueError(f"AdmissionPaths row {row_number}: required value missing")
        if row["program_id"] not in programs:
            raise ValueError(f"AdmissionPaths row {row_number}: unknown program_id {row['program_id']}")
        statements.append(
            "INSERT INTO admission_paths (admission_id, program_id, nama, intake, deadline, requires_supervisor, website_url)\n"
            f"VALUES ({quote_sql(stable_uuid('admission_paths', row['admission_id']))}, {quote_sql(row['program_id'])}, {quote_sql(row['nama'])}, {quote_sql(row['intake'])}, {sql_date(row['deadline'])}, {sql_bool(row['requires_supervisor'])}, {quote_sql(row['website_url'])})\n"
            "ON CONFLICT (admission_id) DO UPDATE\n"
            "SET program_id = EXCLUDED.program_id,\n"
            "    nama = EXCLUDED.nama,\n"
            "    intake = EXCLUDED.intake,\n"
            "    deadline = EXCLUDED.deadline,\n"
            "    requires_supervisor = EXCLUDED.requires_supervisor,\n"
            "    website_url = EXCLUDED.website_url;"
        )
    return statements


def build_admission_funding_sql(workbook) -> list[str]:
    headers, rows = iter_sheet_rows(workbook, "AdmissionFunding")
    require_columns("AdmissionFunding", headers, ["admission_funding_id", "admission_id", "funding_id", "linkage_type"])

    statements = []
    for row_number, row in rows:
        if row["admission_funding_id"] == "" or not has_expected_prefix(row["admission_funding_id"], "AF-"):
            continue
        if row["admission_id"] == "" or row["funding_id"] == "" or row["linkage_type"] == "":
            continue
        statements.append(
            "INSERT INTO admission_funding (admission_funding_id, admission_id, funding_id, linkage_type)\n"
            f"VALUES ({quote_sql(stable_uuid('admission_funding', row['admission_funding_id']))}, {quote_sql(stable_uuid('admission_paths', row['admission_id']))}, {quote_sql(stable_uuid('funding_options', row['funding_id']))}, {quote_sql(row['linkage_type'])})\n"
            "ON CONFLICT (admission_funding_id) DO UPDATE\n"
            "SET admission_id = EXCLUDED.admission_id,\n"
            "    funding_id = EXCLUDED.funding_id,\n"
            "    linkage_type = EXCLUDED.linkage_type;"
        )
    return statements


def build_funding_requirements_sql(workbook, requirement_catalog: dict[str, CatalogRow]) -> list[str]:
    headers, rows = iter_sheet_rows(workbook, "FundingRequirements")
    require_columns("FundingRequirements", headers, ["funding_req_id", "funding_id", "req_catalog_id", "is_required", "sort_order"])

    statements = []
    for row_number, row in rows:
        if row["funding_req_id"] == "" or not has_expected_prefix(row["funding_req_id"], "FR-"):
            continue
        if row["funding_id"] == "" or row["req_catalog_id"] == "" or row["sort_order"] == "":
            continue
        if row["req_catalog_id"] not in requirement_catalog:
            raise ValueError(
                f"FundingRequirements row {row_number}: unknown req_catalog_id {row['req_catalog_id']}"
            )
        req_key = requirement_catalog[row["req_catalog_id"]].key
        statements.append(
            "INSERT INTO funding_requirements (funding_req_id, funding_id, req_catalog_id, is_required, sort_order)\n"
            f"VALUES ({quote_sql(stable_uuid('funding_requirements', row['funding_req_id']))}, {quote_sql(stable_uuid('funding_options', row['funding_id']))}, (SELECT req_catalog_id FROM requirement_catalog WHERE key = {quote_sql(req_key)}), {sql_bool(row['is_required'])}, {sql_int(row['sort_order'])})\n"
            "ON CONFLICT (funding_req_id) DO UPDATE\n"
            "SET funding_id = EXCLUDED.funding_id,\n"
            "    req_catalog_id = EXCLUDED.req_catalog_id,\n"
            "    is_required = EXCLUDED.is_required,\n"
            "    sort_order = EXCLUDED.sort_order;"
        )
    return statements


def build_funding_benefits_sql(workbook, benefit_catalog: dict[str, CatalogRow]) -> list[str]:
    headers, rows = iter_sheet_rows(workbook, "FundingBenefits")
    require_columns("FundingBenefits", headers, ["funding_benefit_id", "funding_id", "benefit_id", "value_text", "sort_order"])

    statements = []
    for row_number, row in rows:
        if row["funding_benefit_id"] == "" or not has_expected_prefix(row["funding_benefit_id"], "FB-"):
            continue
        if row["funding_id"] == "" or row["benefit_id"] == "" or row["value_text"] == "" or row["sort_order"] == "":
            continue
        if row["benefit_id"] not in benefit_catalog:
            raise ValueError(
                f"FundingBenefits row {row_number}: unknown benefit_id {row['benefit_id']}"
            )
        benefit_key = benefit_catalog[row["benefit_id"]].key
        statements.append(
            "INSERT INTO funding_benefits (funding_benefit_id, funding_id, benefit_id, value_text, sort_order)\n"
            f"VALUES ({quote_sql(stable_uuid('funding_benefits', row['funding_benefit_id']))}, {quote_sql(stable_uuid('funding_options', row['funding_id']))}, (SELECT benefit_id FROM benefit_catalog WHERE key = {quote_sql(benefit_key)}), {quote_sql(row['value_text'])}, {sql_int(row['sort_order'])})\n"
            "ON CONFLICT (funding_benefit_id) DO UPDATE\n"
            "SET funding_id = EXCLUDED.funding_id,\n"
            "    benefit_id = EXCLUDED.benefit_id,\n"
            "    value_text = EXCLUDED.value_text,\n"
            "    sort_order = EXCLUDED.sort_order;"
        )
    return statements


def build_sql(workbook_path: Path) -> str:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    countries = collect_countries(workbook)
    universities = collect_universities(workbook, countries)
    programs = collect_programs(workbook)
    requirement_catalog = collect_catalog_rows(workbook, "RequirementCatalog", "req_catalog_id")
    benefit_catalog = collect_catalog_rows(workbook, "BenefitCatalog", "benefit_id")

    sections = [
        "-- Generated from Excel workbook for GlobalMatch/DreamTracker master data.",
        "BEGIN;",
    ]

    for group in (
        build_countries_sql(countries),
        build_universities_sql(universities),
        build_programs_sql(programs, universities),
        build_requirement_catalog_sql(workbook),
        build_benefit_catalog_sql(workbook),
        build_funding_options_sql(workbook),
        build_admission_paths_sql(workbook, programs),
        build_admission_funding_sql(workbook),
        build_funding_requirements_sql(workbook, requirement_catalog),
        build_funding_benefits_sql(workbook, benefit_catalog),
    ):
        sections.extend(group)
        sections.append("")

    sections.append("COMMIT;")
    return "\n".join(sections).strip() + "\n"


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate SQL seed statements from the Boundless data migration Excel workbook."
    )
    parser.add_argument("workbook", help="Path to the .xlsx workbook")
    parser.add_argument("-o", "--output", help="Write SQL to a file instead of stdout")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        print(f"workbook not found: {workbook_path}", file=sys.stderr)
        return 1

    try:
        sql = build_sql(workbook_path)
    except Exception as exc:
        print(f"failed to generate SQL: {exc}", file=sys.stderr)
        return 1

    if args.output:
        output_path = Path(args.output)
        output_path.write_text(sql, encoding="utf-8")
    else:
        sys.stdout.write(sql)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
